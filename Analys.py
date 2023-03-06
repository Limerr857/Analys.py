#--------------------
#      Settings     
#--------------------

# Maximum acceptable distance in pixels between two point sources in image 28 and 29 
# where they are still counted as the same star.
max_dist = 3.0

# Whether the program should output a histogram plot or a scatter plot of the data
plot_histogram = False

# The bin size for the histogram plot
histogram_bin = 10

# This value is the average of PSF to aperture magnitude corrections for five bright, isolated stars
Ko1_28 = 0.1516
Ko1_29 = 0.1416

# This value is the share of the stars total magnitude caught by the 0.5 " aperture
Ko2_28 = 0.916
Ko2_29 = 0.914

# This value is the PSF magnitude of a star outputting exactly one electron per second 
# when its photons hit the CCD sensor 
Ko3_28 = 18.371
Ko3_29 = 17.875

# This value is the apparent magnitude of a star outputting exactly one electron per second 
# when its photons hit the CCD sensor
Ko4_28 = 26.403
Ko4_29 = 25.507



#--------------------
#      Program     
#--------------------

# Import required libraries
import matplotlib.pyplot as plt
import openpyxl as pxl
import math

# Import data from spreadsheet files
book_28 = pxl.load_workbook("28.xlsx")
sh_28 = book_28.active
book_29 = pxl.load_workbook("29.xlsx")
sh_29 = book_29.active

# Create lists to store data obtained from the spreadsheet files
x_y_mag_28 = []
x_y_mag_29 = []



### Transfer the data from the spreadsheet files to the lists ###

# Create a smaller list for every identified star in 28.xlsx and 
# append it to the larger list. The smaller list will later be used to store 
# the x position, y position and magnitude of each star
for i in range(sh_28.max_row):
    x_y_mag_28.append([0,0,0])

# Fill the list of lists with the x position, 
# y position and magnitude of each star
# Iterate over the cells in 28.xlsx that contain the data needed for analysis
for row in sh_28.iter_rows(min_row=1, min_col=2, max_col=4, max_row=sh_28.max_row):
    for cell in row:
        # Fill one specific element of one specific list 
        # inside the list of lists with the correct value
        x_y_mag_28[cell.row-1][cell.column-2] = cell.value

# Create a smaller list for every identified star in 29.xlsx and 
# append it to the larger list. The smaller list will later be used to store 
# the x position, y position and magnitude of each star
for i in range(sh_29.max_row):
    x_y_mag_29.append([0,0,0])

# Fill the list of lists with the x position, 
# y position and magnitude of each star
# Iterate over the cells in 29.xlsx that contain the data needed for analysis
for row in sh_29.iter_rows(min_row=1, min_col=2, max_col=4, max_row=sh_29.max_row):
    for cell in row:
        # Fill one specific element of one specific list 
        # inside the list of lists with the correct value
        x_y_mag_29[cell.row-1][cell.column-2] = cell.value



### Correct the star magnitudes ###

for star in x_y_mag_28:
    
    # Convert from PSF magnitude to aperture 0.5 " magnitude
    star[2] += Ko1_28 

    # Convert from aperture 0.5 " magnitude to aperture infinite magnitude
    star[2] *= Ko2_28 
    
    # Subtract the zero point for the PSF
    star[2] -= Ko3_28
    
    # Add the VEGAMAG zero point for ACS/WFC at the appropriate date
    star[2] += Ko4_28

for star in x_y_mag_29:
    
    # Convert from PSF magnitude to aperture 0.5 " magnitude
    star[2] += Ko1_29

    # Convert from aperture 0.5 " magnitude to aperture infinite magnitude
    star[2] *= Ko2_29
    
    # Subtract the zero point for the PSF
    star[2] -= Ko3_29
    
    # Add the VEGAMAG zero point for ACS/WFC at the appropriate date
    star[2] += Ko4_29



### Match stars between the two images ###

# Create a list to hold the matched magnitudes
mag28_mag29 = []

# Iterate over every star in image 28 and find the closest matching star in image 29,
# so that every star's magnitude in each filter can be used in analysis
for star28 in x_y_mag_28:
    
    # Set temporary values which will be overridden later
    closest_range = 9999
    mag29_for_best_fit = -1
    
    for star29 in x_y_mag_29:
        # Calculate the distance between a star in image 28 and a star in image 29
        dist = math.sqrt( (star28[0]-star29[0])**2 + (star28[1]-star29[1])**2 )
        
        # If the distance is the closest so far
        if (dist < closest_range):
            
            # Set a new closest distance
            closest_range = dist
            
            # And save the magnitude of the star in image 29
            mag29_for_best_fit = star29[2]
            
    # If the closest star is not further away than the mximum distance,
    # keep it for analysis
    if (closest_range <= max_dist):
        mag28_mag29.append( [ round(star28[2], 3), round(mag29_for_best_fit, 3) ] )



### Plot the data ###

# Create the lists containing the magnitudes...
mag29 = [] # ... for each star in image 29
magIV = [] # ... for each star in image 28 
           # subtracted by the magnitude in image 29

# Fill the lists mentioned above
for star in mag28_mag29:
    mag29.append(star[1])
    magIV.append(star[0]-star[1])

# Plot the data in matplotlib, user decides whether 
# histogram plot or scatter plot should be used
if plot_histogram == False:
    
    # Plot the scatterplot with the I-V magnitude on the X-axis
    # and the magnitude in image 29 on the Y-axis. 
    # s=1 makes the point size small enough to distinguish individual star indexes 
    plt.scatter(magIV,mag29,s=1)
    
    # These axis values can be changed, but matplotlib also allows the user 
    # to pan and zoom the scatterplot
    plt.axis([-1,2.5,31,26])
    
else:
    
    # Plot the histogram of the magnitudes in image 29 with bin 10
    plt.hist(mag29, histogram_bin)

# Show the recently created diagram to the user
plt.show()